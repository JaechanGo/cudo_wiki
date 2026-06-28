"""xlsx 수당표/전결표 셀 추출 (plan §6, openpyxl).

두 진입점:
- ``extract_excel_cells``: 첫 시트의 헤더 행을 인식해 각 데이터 행을 **authority cells dict**
  (파트1 인계 스키마)로 매핑한다. 산출 dict 는 그대로 ``authority_parser.parse_authority_matrix``
  입력으로 호환된다(business_item·action_type·amount_expr·consulter_roles·...).
- ``extract_excel_text``: 셀 전체를 plain text 로 덤프(attachment.extracted_text 용).

LLM 미사용 — 헤더명 매핑 + 결정론 규칙(인용 결정론, NFR). 알 수 없는 헤더 열은 무시한다.
"""

from __future__ import annotations

import io
import re

import openpyxl

# 헤더명(한국어/영문) → authority cells dict 키. 부분 일치(헤더에 키워드 포함)로 매핑.
_HEADER_MAP: tuple[tuple[tuple[str, ...], str], ...] = (
    (("대분류", "분류", "category"), "business_category"),
    (("업무항목", "업무", "항목", "item"), "business_item"),
    (("결재구분", "전결구분", "구분", "결재", "action"), "action_type"),
    (("금액", "amount"), "amount_expr"),
    (("전결권자", "결재자", "전결자", "approver"), "approver_role"),
    (("합의", "협의", "consult"), "consulter_roles"),
    (("비고", "조건", "note"), "condition_note"),
    (("행", "row"), "matrix_row_label"),
    (("열", "col"), "matrix_col_label"),
)

# consulter_roles 분리자: 슬래시·콤마·중점·세미콜론.
_ROLE_SPLIT_RE = re.compile(r"[/,·;]")


def _match_field(header: str) -> str | None:
    """헤더 셀 문자열 → cells dict 키(부분 일치). 매칭 없으면 None."""
    h = (header or "").strip()
    if not h:
        return None
    for keywords, field in _HEADER_MAP:
        if any(kw in h for kw in keywords):
            return field
    return None


def _load_first_sheet(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return wb[wb.sheetnames[0]]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_excel_cells(content: bytes) -> list[dict]:
    """전결표 xlsx → authority cells dict 목록.

    첫 행을 헤더로 보고 ``_HEADER_MAP`` 으로 필드명을 해소한다. 이후 각 데이터 행을 dict 로
    매핑하며, ``consulter_roles`` 는 구분자(/,·;)로 분할, ``order_seq`` 는 데이터 행 인덱스(0..N).
    전부 빈 행은 건너뛴다.
    """
    ws = _load_first_sheet(content)
    rows = ws.iter_rows(values_only=True)

    header_row = next(rows, None)
    if header_row is None:
        return []
    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        field = _match_field(_cell_str(cell))
        if field is not None:
            col_field[idx] = field

    cells: list[dict] = []
    seq = 0
    for raw in rows:
        values = [_cell_str(v) for v in raw]
        if not any(values):
            continue
        record: dict = {}
        for idx, field in col_field.items():
            val = values[idx] if idx < len(values) else ""
            if not val:
                continue
            if field == "consulter_roles":
                record[field] = [p.strip() for p in _ROLE_SPLIT_RE.split(val) if p.strip()]
            else:
                record[field] = val
        if not record.get("business_item"):
            continue
        record["order_seq"] = seq
        seq += 1
        cells.append(record)
    return cells


def extract_excel_text(content: bytes) -> str:
    """셀 전체를 탭/개행으로 이은 plain text 덤프(검색 본문·디버그용)."""
    ws = _load_first_sheet(content)
    lines: list[str] = []
    for raw in ws.iter_rows(values_only=True):
        values = [_cell_str(v) for v in raw]
        if any(values):
            lines.append("\t".join(values))
    return "\n".join(lines)
